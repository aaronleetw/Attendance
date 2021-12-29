from functions import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io

center = Alignment(horizontal="center", vertical="center")
std_font = Font(name="Calibri", size=13)
side = Side(border_style='thin')
border = Border(left=side, right=side, top=side, bottom=side)
bold_bottom = Border(left=side, right=side, top=side, bottom=Side(border_style='medium', color='FF000000'))


def create_period_sheets(workbook, class_code):
    ws = workbook.create_sheet(class_code[0] + class_code[1])
    ws.merge_cells('A1:F1')
    ws['A1'] = '台北市私立復興實驗高級中學班級課表'
    ws['A1'].font = Font(name="DFKai-SB", size=15, bold=True)
    ws['A1'].alignment = center
    # loop over A:F
    for i in range(0, 6):
        ws[str(chr(ord('A') + i)) + '1'].border = border
    ws['G1'] = class_code[0] + class_code[1]
    ws['G1'].font = Font(name='Courier New', size=20, bold=True)
    ws['G1'].alignment = center
    ws['G1'].border = border

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 25
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 10

    ws.merge_cells('A2:B2')
    ws['A2'] = '時間'
    ws['C2'] = '星期一'
    ws['D2'] = '星期二'
    ws['E2'] = '星期三'
    ws['F2'] = '星期四'
    ws['G2'] = '星期五'

    # loop over A2:G2
    for i in range(1, 8):
        ws.cell(row=2, column=i).font = Font(size=14, bold=True)
        ws.cell(row=2, column=i).alignment = center
        ws.cell(row=2, column=i).border = border

    # loop over C:G
    for i in range(2, 8):
        ws.column_dimensions[str(chr(ord('A') + i))].width = 13

    # get data
    db = refresh_db()
    cursor = db.cursor()
    cursor.execute("SELECT dow,period,subject,teacher FROM schedule WHERE grade=%s AND class_=%s",
                   (class_code[0], class_code[1]))
    sql = cursor.fetchall()
    data = {}
    subject_teacher = {}
    # loop over data
    for i in sql:
        if i[0] not in data:
            data[i[0]] = {}
        data[i[0]][i[1]] = {
            'subject': i[2],
            'teacher': i[3]
        }
        if i[2] != 'GP' and i[2] != '--' and i[3] != '--' and i[2] not in subject_teacher:
            subject_teacher[i[2]] = i[3]

    periods = ['m', '1', '2', '3', '4', 'n', '5', '6', '7', '8', '9']
    times = {
        'm': ['7:30', '8:10'],
        '1': ['8:20', '9:05'],
        '2': ['9:15', '10:00'],
        '3': ['10:10', '10:55'],
        '4': ['11:05', '11:50'],
        'n': ['11:50', '13:05'],
        '5': ['13:15', '14:00'],
        '6': ['14:10', '14:55'],
        '7': ['15:05', '15:50'],
        '8': ['15:55', '16:40'],
        '9': ['16:45', '17:30']
    }
    curr = 3
    for p in periods:
        ws.merge_cells('A' + str(curr) + ':A' + str(curr + 1))
        ws.row_dimensions[curr].height = 20
        ws.row_dimensions[curr + 1].height = 20
        ws['A' + str(curr)] = p
        ws['A' + str(curr)].font = std_font
        ws['A' + str(curr)].alignment = center
        ws['A' + str(curr)].border = border
        ws['A' + str(curr + 1)].border = border
        ws['B' + str(curr)] = times[p][0]
        ws['B' + str(curr)].font = std_font
        ws['B' + str(curr)].alignment = center
        ws['B' + str(curr)].border = border
        ws['B' + str(curr + 1)] = times[p][1]
        ws['B' + str(curr + 1)].font = std_font
        ws['B' + str(curr + 1)].alignment = center
        ws['B' + str(curr + 1)].border = border

        if p == 'm' or p == 'n':
            ws.merge_cells('C' + str(curr) + ':G' + str(curr + 1))
            for i in range(1, 6):
                ws[chr(ord('C') + i - 1) + str(curr)].font = std_font
                ws[chr(ord('C') + i - 1) + str(curr)].alignment = center
                ws[chr(ord('C') + i - 1) + str(curr)].border = border
                ws[chr(ord('C') + i - 1) + str(curr + 1)].border = border
            if p == 'm':
                ws['C' + str(curr)] = '早自習'
            else:
                ws['C' + str(curr)] = '午餐 / 午休'
        else:
            for i in range(1, 6):
                ws.merge_cells(chr(ord('C') + i - 1) + str(curr) + ':' + chr(ord('C') + i - 1) + str(curr + 1))
                ws[chr(ord('C') + i - 1) + str(curr)].font = std_font
                ws[chr(ord('C') + i - 1) + str(curr)].alignment = center
                ws[chr(ord('C') + i - 1) + str(curr)].border = border
                ws[chr(ord('C') + i - 1) + str(curr + 1)].border = border
                if i in data:
                    if p in data[i]:
                        ws[chr(ord('C') + i - 1) + str(curr)] = (data[i][p]['subject'] if data[i][p]['subject'] != 'GP'
                                                                                          and data[i][p][
                                                                                              'subject'] != '--' else '' if
                        data[i][p]['subject'] == '--' else data[i][p]['teacher'])
        curr += 2
    ws.merge_cells('A26:G26')
    ws['A26'] = '科任老師一覽表'
    ws['A26'].font = Font(size=14, bold=True)
    ws['A26'].alignment = center
    ws.row_dimensions[26].height = 20
    # loop over A26:G26
    for i in range(0, 7):
        ws[chr(ord('A') + i) + '26'].border = border
    curr = 0
    for i in subject_teacher:
        if (curr % 3) == 0:
            pos = ['A', 'C']
        elif (curr % 3) == 1:
            pos = ['D', 'E']
        else:
            pos = ['F', 'G']
        loc = str(27 + int(curr / 3))
        ws.merge_cells(pos[0] + loc + ':' + pos[1] + loc)
        ws[pos[0] + loc].font = std_font
        ws[pos[0] + loc].alignment = center
        ws[pos[0] + loc].border = border
        ws[pos[0] + loc] = i + ': ' + subject_teacher[i]
        for j in range(ord(pos[0]), ord(pos[1]) + 1):
            ws[chr(j) + loc].border = border
        ws.row_dimensions[curr + 27].height = 20
        curr += 1
    return workbook


def create_student_list(workbook, class_code):
    ws = workbook.create_sheet(class_code[0] + class_code[1])
    ws.merge_cells('A1:J1')
    ws['A1'] = '台北市私立復興實驗高級中學學生名單'
    ws['A1'].font = Font(name="DFKai-SB", size=15, bold=True)
    ws['A1'].alignment = center
    # loop over A:J
    for i in range(0, 11):
        ws[str(chr(ord('A') + i)) + '1'].border = border
    ws.merge_cells('K1:L1')
    ws['K1'] = class_code[0] + class_code[1]
    ws['K1'].font = Font(name='Courier New', size=20, bold=True)
    ws['K1'].alignment = center
    ws['K1'].border = border
    ws['L1'].border = border
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 12
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20

    ws['A2'] = '#'
    ws['A2'].font = Font(name="Calibri", size=13, bold=True)
    ws['A2'].alignment = center
    ws['A2'].border = bold_bottom
    ws['B2'] = '姓名'
    ws['B2'].font = Font(name="DFKai-SB", size=13, bold=True)
    ws['B2'].alignment = center
    ws['B2'].border = bold_bottom
    ws['C2'] = 'Name'
    ws['C2'].font = Font(name="Calibri", size=13, bold=True)
    ws['C2'].alignment = center
    ws['C2'].border = bold_bottom
    for i in range(3, 12):
        ws[str(chr(ord('A') + i)) + '2'].border = bold_bottom
        ws.column_dimensions[str(chr(ord('A') + i))].width = 5.8
    db = refresh_db()
    cursor = db.cursor()
    cursor.execute('SELECT num,name,ename FROM students WHERE grade=%s AND class_=%s ORDER BY num ASC',
                   (class_code[0], class_code[1]))
    data = cursor.fetchall()
    last = data[-1][0]
    delcnt = 0
    for i in range(0, last):
        ws['A' + str(3 + i)] = i + 1
        ws['A' + str(3 + i)].font = std_font
        ws['A' + str(3 + i)].alignment = center
        ws['B' + str(3 + i)] = data[i - delcnt][1] if data[i - delcnt][0] == i + 1 else ''
        ws['B' + str(3 + i)].font = Font(name="DFKai-SB", size=14)
        ws['B' + str(3 + i)].alignment = center
        ws['C' + str(3 + i)] = data[i - delcnt][2] if data[i - delcnt][0] == i + 1 else ''
        ws['C' + str(3 + i)].font = std_font
        ws['C' + str(3 + i)].alignment = center
        ws.row_dimensions[3 + i].height = 19
        for j in range(0, 12):
            ws[str(chr(ord('A') + j)) + str(3 + i)].border = bold_bottom if (i + 1) % 5 == 0 else border
        if data[i - delcnt][0] != i + 1:
            delcnt += 1
    return workbook


def create_group_student_list(workbook, cclass, data):
    class_code = [cclass['category'], cclass['class_id']]
    ws = workbook.create_sheet(class_code[0] + '.' + class_code[1])
    ws.merge_cells('A1:H1')
    ws['A1'] = '台北市私立復興實驗高級中學分組課學生名單'
    ws['A1'].font = Font(name="DFKai-SB", size=15, bold=True)
    ws['A1'].alignment = center
    # loop over A:I
    for i in range(0, 10):
        ws[str(chr(ord('A') + i)) + '1'].border = border
    ws.merge_cells('I1:L1')
    ws['I1'] = class_code[0] + class_code[1]
    ws['I1'].font = Font(name='Calibri', size=15, bold=True)
    ws['I1'].alignment = center
    ws['I1'].border = border
    ws['J1'].border = border
    ws['K1'].border = border
    ws['L1'].border = border
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 12
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20

    ws['A2'] = '班級'
    ws['A2'].font = Font(name="Calibri", size=13, bold=True)
    ws['A2'].alignment = center
    ws['A2'].border = bold_bottom
    ws['B2'] = '#'
    ws['B2'].font = Font(name="Calibri", size=13, bold=True)
    ws['B2'].alignment = center
    ws['B2'].border = bold_bottom
    ws['C2'] = '姓名'
    ws['C2'].font = Font(name="DFKai-SB", size=13, bold=True)
    ws['C2'].alignment = center
    ws['C2'].border = bold_bottom
    ws['D2'] = 'Name'
    ws['D2'].font = Font(name="Calibri", size=13, bold=True)
    ws['D2'].alignment = center
    ws['D2'].border = bold_bottom
    for i in range(4, 12):
        ws[str(chr(ord('A') + i)) + '2'].border = bold_bottom
        ws.column_dimensions[str(chr(ord('A') + i))].width = 5.8

    cnt = 0
    for i in data:
        ws['A' + str(3 + cnt)] = str(i[0]) + str(i[1])
        ws['A' + str(3 + cnt)].font = std_font
        ws['A' + str(3 + cnt)].alignment = center
        ws['B' + str(3 + cnt)] = i[2]
        ws['B' + str(3 + cnt)].font = std_font
        ws['B' + str(3 + cnt)].alignment = center
        ws['C' + str(3 + cnt)] = i[3]
        ws['C' + str(3 + cnt)].font = Font(name="DFKai-SB", size=14)
        ws['C' + str(3 + cnt)].alignment = center
        ws['D' + str(3 + cnt)] = i[4]
        ws['D' + str(3 + cnt)].font = std_font
        ws['D' + str(3 + cnt)].alignment = center
        ws.row_dimensions[3 + cnt].height = 19
        for j in range(0, 12):
            ws[str(chr(ord('A') + j)) + str(3 + cnt)].border = bold_bottom if (cnt + 1) % 5 == 0 else border
        cnt += 1
    return workbook


def create_teacher_periods(workbook, teacher_name, orig_username=''):
    ws = workbook.create_sheet(teacher_name)
    ws.merge_cells('A1:E1')
    ws['A1'] = '台北市私立復興實驗高級中學科任老師課表'
    ws['A1'].font = Font(name="DFKai-SB", size=15, bold=True)
    ws['A1'].alignment = center
    # loop over A:E
    for i in range(0, 5):
        ws[str(chr(ord('A') + i)) + '1'].border = border
    ws.merge_cells('F1:G1')
    ws['F1'] = teacher_name + " 老師"
    ws['F1'].font = Font(name='Calibri', size=15, bold=True)
    ws['F1'].alignment = center
    ws['F1'].border = border
    ws['G1'].border = border

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 25
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 10

    ws.merge_cells('A2:B2')
    ws['A2'] = '時間'
    ws['C2'] = '星期一'
    ws['D2'] = '星期二'
    ws['E2'] = '星期三'
    ws['F2'] = '星期四'
    ws['G2'] = '星期五'

    # loop over A2:G2
    for i in range(1, 8):
        ws.cell(row=2, column=i).font = Font(size=14, bold=True)
        ws.cell(row=2, column=i).alignment = center
        ws.cell(row=2, column=i).border = border

    # loop over C:G
    for i in range(2, 8):
        ws.column_dimensions[str(chr(ord('A') + i))].width = 13

    # get data
    data = {}
    db = refresh_db()
    cursor = db.cursor()
    if orig_username is not '':
        cursor.execute('SELECT category,subclass FROM gpclasses WHERE accs LIKE %s', ('%' + orig_username + '%',))
        gp_sql = cursor.fetchall()
        for i in gp_sql:
            cursor.execute('SELECT dow,period FROM schedule WHERE teacher=%s', (i[0],))
            tmp_sql = cursor.fetchall()
            for j in tmp_sql:
                if j[0] not in data:
                    data[j[0]] = {}
                data[j[0]][j[1]] = {
                    'subject': i[0],
                    'class': i[1]
                }
    cursor.execute("SELECT dow,period,subject,grade,class_ FROM schedule WHERE teacher=%s", (teacher_name,))
    sql = cursor.fetchall()
    # loop over data
    for i in sql:
        if i[0] not in data:
            data[i[0]] = {}
        data[i[0]][i[1]] = {
            'subject': i[2],
            'class': str(i[3]) + str(i[4])
        }

    periods = ['m', '1', '2', '3', '4', 'n', '5', '6', '7', '8', '9']
    times = {
        'm': ['7:30', '8:10'],
        '1': ['8:20', '9:05'],
        '2': ['9:15', '10:00'],
        '3': ['10:10', '10:55'],
        '4': ['11:05', '11:50'],
        'n': ['11:50', '13:05'],
        '5': ['13:15', '14:00'],
        '6': ['14:10', '14:55'],
        '7': ['15:05', '15:50'],
        '8': ['15:55', '16:40'],
        '9': ['16:45', '17:30']
    }
    curr = 3
    for p in periods:
        ws.merge_cells('A' + str(curr) + ':A' + str(curr + 1))
        ws.row_dimensions[curr].height = 20
        ws.row_dimensions[curr + 1].height = 20
        ws['A' + str(curr)] = p
        ws['A' + str(curr)].font = std_font
        ws['A' + str(curr)].alignment = center
        ws['A' + str(curr)].border = border
        ws['A' + str(curr + 1)].border = border
        ws['B' + str(curr)] = times[p][0]
        ws['B' + str(curr)].font = std_font
        ws['B' + str(curr)].alignment = center
        ws['B' + str(curr)].border = border
        ws['B' + str(curr + 1)] = times[p][1]
        ws['B' + str(curr + 1)].font = std_font
        ws['B' + str(curr + 1)].alignment = center
        ws['B' + str(curr + 1)].border = border

        if p == 'm' or p == 'n':
            ws.merge_cells('C' + str(curr) + ':G' + str(curr + 1))
            for i in range(1, 6):
                ws[chr(ord('C') + i - 1) + str(curr)].font = std_font
                ws[chr(ord('C') + i - 1) + str(curr)].alignment = center
                ws[chr(ord('C') + i - 1) + str(curr)].border = border
                ws[chr(ord('C') + i - 1) + str(curr + 1)].border = border
            if p == 'm':
                ws['C' + str(curr)] = '早自習'
            else:
                ws['C' + str(curr)] = '午餐 / 午休'
        else:
            for i in range(1, 6):
                ws.merge_cells(chr(ord('C') + i - 1) + str(curr) + ':' + chr(ord('C') + i - 1) + str(curr + 1))
                ws[chr(ord('C') + i - 1) + str(curr)].font = std_font
                ws[chr(ord('C') + i - 1) + str(curr)].border = border
                ws[chr(ord('C') + i - 1) + str(curr + 1)].border = border
                ws[chr(ord('C') + i - 1) + str(curr)].alignment = center + Alignment(wrapText=True)
                if i in data:
                    if p in data[i]:
                        ws[chr(ord('C') + i - 1) + str(curr)] = (
                            data[i][p]['subject'] + '\n' + data[i][p]['class'] if data[i][p]['subject'] != 'GP'
                                                                                  and data[i][p][
                                                                                      'subject'] != '--' else '' if
                            data[i][p]['subject'] == '--' else data[i][p]['teacher'])
        curr += 2
    return workbook
